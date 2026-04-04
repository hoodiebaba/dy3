import os
import re
import pandas as pd
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo


class GSAuditReport:
    def __init__(self, usid, gs_admin=False):
        self.usid = usid
        self.audit_file = os.path.join(self.usid.outdir, F'{self.usid.site_id}_GS_Report.xlsx')
        df_temp = self.usid.df_report.copy(deep=True)
        
        if len(df_temp.index) > 0:
            wb_final = pd.ExcelWriter(self.audit_file, engine='openpyxl')
            df_temp = df_temp[['Site', 'MO', 'MOC', 'Type', 'Parameter', 'CurrentValue', 'GSValue', 'InitialValue', 'Permission', 'Suffix', 'flag']]
            df_temp.MO = df_temp.MO.str.extract(r'.*,ManagedElement=\w+,(.*)').squeeze()
            df_temp.to_excel(wb_final, 'Audit_Report', index=False, header=True)
            self.report_create_formating(wb_final.sheets['Audit_Report'], 'Audit_Report')
            wb_final.close()
        
        # Relations Table for Admin User Only
        if not gs_admin: return
        if len(df_temp.index) > 0:
            df_gs_delta = df_temp.loc[(~df_temp.flag) & (df_temp.GSValue != 'GS_Error')]
            df_gs_error = df_temp.loc[df_temp.GSValue == 'GS_Error']
            if len(df_gs_delta.index) > 0 or len(df_gs_error.index) > 0:
                wb_final = pd.ExcelWriter(os.path.join(self.usid.outdir, F'{self.usid.site_id}_GS_Delta_Error.xlsx'), engine='openpyxl')
                if len(df_gs_delta.index) > 0:
                    df_gs_delta.to_excel(wb_final, 'Delta', index=False, header=True)
                    self.report_create_formating(wb_final.sheets['Delta'], 'Delta')
                if len(df_gs_error.index) > 0:
                    df_gs_error.to_excel(wb_final, 'Error', index=False, header=True)
                    self.report_create_formating(wb_final.sheets['Error'], 'Error')
                wb_final.close()
            
        wb_para = pd.ExcelWriter(os.path.join(self.usid.outdir, F'{self.usid.site_id}_Logic.xlsx'), engine='openpyxl')
        para_dict = {'USID': 'para', 'SITE': 'sites', 'NR_CELL': 'nr_cells', 'LTE_CELL': 'cells',
                     'EARFCN': 'earfcn', 'UARFCN': 'uarfcn', 'ARFCNNR': 'ssbfreq'}
        df_temp = pd.DataFrame([])
        for sheet in para_dict.keys():
            if sheet in ['USID']:
                df_temp = pd.DataFrame([self.usid.param_dict.get(para_dict[sheet], {})])
            elif sheet in ['SITE']:
                df_temp = pd.DataFrame([self.usid.param_dict.get(para_dict[sheet]).get(site).get('para')
                                        for site in self.usid.param_dict.get(para_dict[sheet], {}).keys()])
            elif sheet in ['NR_CELL', 'LTE_CELL']:
                temp_list = []
                for site in self.usid.param_dict.get('sites').keys():
                    for cell in self.usid.param_dict.get('sites').get(site).get(para_dict[sheet], {}).keys():
                        temp_list.append(self.usid.param_dict.get('sites').get(site).get(para_dict[sheet]).get(cell))
                df_temp = pd.DataFrame(temp_list)
            elif sheet in ['EARFCN', 'UARFCN', 'ARFCNNR']:
                temp_list = []
                for key in self.usid.param_dict.get(para_dict[sheet], {}).keys():
                    tmp = self.usid.param_dict.get(para_dict[sheet], {}).get(key)
                    tmp['freq'] = key
                    temp_list.append(tmp)
                df_temp = pd.DataFrame(temp_list)
            if len(df_temp.index) > 0:
                df_temp.columns = df_temp.columns.astype(str)
                df_temp = df_temp.add_prefix('_')
                df_temp.to_excel(wb_para, sheet, index=False, header=True)
                self.excell_table_formating(wb_para.sheets[sheet], sheet)
        # if save_flag: wb_para.save()
        wb_para.close()
        
        wb_para = pd.ExcelWriter(os.path.join(self.usid.outdir, F'{self.usid.site_id}_Relations.xlsx'), engine='openpyxl')
        para_dict = {'lte_freq': self.usid.df_lte_freq, 'lte_rel': self.usid.df_lte_rel, 'lte_crel': self.usid.df_lte_crel,
                     'lte_umts_freq': self.usid.df_lte_umts_freq, 'lte_umts_rel': self.usid.df_lte_umts_rel,
                     'lte_nr_freq': self.usid.df_lte_nr_freq, 'lte_nr_rel': self.usid.df_lte_nr_rel, 'lte_nr_crel': self.usid.df_lte_nr_crel,
                     'nr_freq': self.usid.df_nr_freq, 'nr_rel': self.usid.df_nr_rel, 'nr_crel': self.usid.df_nr_crel}
        for sheet in para_dict.keys():
            df_temp = para_dict[sheet].copy(deep=True)
            if len(df_temp.index) > 0:
                df_temp.columns = df_temp.columns.astype(str)
                df_temp.to_excel(wb_para, sheet, index=False, header=True)
                self.excell_table_formating(wb_para.sheets[sheet], sheet)
        # if save_flag: wb_para.save()
        wb_para.close()
        
    @staticmethod
    def report_create_formating(work_sheet, sheet_name):
        """
            This method fetches the worksheet dimensions and returns the dimensions plus a dictionary that contains a
            mapping of the sheet column names to their cell coordinates
        """
        dims = work_sheet.calculate_dimension()
        matched = re.match(r"([A-Z]+)[\d]+:([A-Z]+)[\d]+", dims)
        column_ref = ":".join([str_ + '1' for str_ in matched.groups()])
        col_coords = dict([(cell_.value, cell_.coordinate) for cell_ in work_sheet[column_ref][0]])
        
        # Add Red Colour for error flag
        dxf_error = DifferentialStyle(fill=PatternFill("solid", bgColor="FF0000"), font=Font(b=False, color="FFFFFF"))
        r_error = Rule(type="expression", dxf=dxf_error, stopIfTrue=True)
        r_error.formula = [F"${col_coords['GSValue']}=\"error\""]
        work_sheet.conditional_formatting.add(dims, r_error)
        
        # If 'flag' is False & Permission is Global or GlobalOptional, then it should be filled Red
        dxf_global = DifferentialStyle(fill=PatternFill("solid", bgColor="F47174"))
        r_global = Rule(type="expression", dxf=dxf_global, stopIfTrue=True)
        r_global.formula = [F"AND(OR(${col_coords['Permission']}=\"Global\", ${col_coords['Permission']}=\"GlobalOptional\", $"
                            F"{col_coords['Permission']}=\"GlobalOptional \"),${col_coords['flag']}=FALSE)"]
        work_sheet.conditional_formatting.add(dims, r_global)
        
        # If 'Flag' is False & Permission is Not Auditable, then it should be filled Grey
        dxf_not_auditable = DifferentialStyle(fill=PatternFill("solid", bgColor="B7825F"))
        r_not_auditable = Rule(type="expression", dxf=dxf_not_auditable, stopIfTrue=True)
        r_not_auditable.formula = [F"AND(${col_coords['Permission']}=\"Not Auditable\",${col_coords['flag']}=FALSE)"]
        work_sheet.conditional_formatting.add(dims, r_not_auditable)
        
        # If 'Flag' is False or True & Permission is Local, then it should be filled Light Blue
        dxf_local = DifferentialStyle(fill=PatternFill("solid", bgColor="93CAED"))
        r_local = Rule(type="expression", dxf=dxf_local, stopIfTrue=True)
        r_local.formula = [F"AND(${col_coords['Permission']}=\"Local\",${col_coords['flag']}=FALSE)"]
        work_sheet.conditional_formatting.add(dims, r_local)
        
        # If 'Flag' is False or True & Permission is Internal, then it should be filled Orange
        dxf_internal = DifferentialStyle(fill=PatternFill("solid", bgColor="F5CA7B"))
        r_internal = Rule(type="expression", dxf=dxf_internal, stopIfTrue=True)
        r_internal.formula = [F"AND(${col_coords['Permission']}=\"Internal\",${col_coords['flag']}=FALSE)"]
        work_sheet.conditional_formatting.add(dims, r_internal)
        
        """
        This method adds a table to the excel sheet. There should be no null column names to prevent when the table is created.
        This is to prevent the excel sheet from getting corrupted.
        """
        tab = Table(displayName=sheet_name + "_Table", name=sheet_name + "_Table", ref=dims)
        # TableStyleLight15
        style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=True, showLastColumn=True, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        work_sheet.add_table(tab)
    
    @staticmethod
    def excell_table_formating(work_sheet, sheet_name):
        tab = Table(displayName=sheet_name + "_Table", name=sheet_name + "_Table", ref=work_sheet.calculate_dimension())
        style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=True, showLastColumn=True, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        work_sheet.add_table(tab)
